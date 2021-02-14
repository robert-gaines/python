#!/usr/bin/env python3

from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import socket
import os

socket.setdefaulttimeout(.5)

def ResolveAddr(addr):
    #
    try:
        hostname = socket.gethostbyaddr(addr)
        if(hostname):
            hostname = hostname[0]
    except:
        hostname = 'Unidentified'
    return hostname

def ReadHostAddrs(fileObj):
    workbook = fileObj
    for w in workbook:
        for row in w.iter_rows():
            for cell in row:
                w.cell(row=1,column=2).value = 'Hostname'
                w.column_dimensions[get_column_letter(2)].width = 30.0
                if(cell.internal_value != "IP Address" and cell.internal_value != 'Hostname'):
                    #column_index = 'B'
                    column_index = 2
                    row_index    = int(cell.row)
                    #print("Writing: ",w,index)
                    print("[*] Attempting to resolve: ",cell.internal_value)
                    hostname = ResolveAddr(cell.internal_value)
                    print("[*] Result: ",hostname)
                    w.cell(row=row_index,column=column_index).value = hostname
            
def main():
    print("[*] Multi-VLAN DNS Resolution Script [*]")
    print("Listing files in the current working directory...")
    for file in os.listdir():
        print(file)
    fileName   = input("[+] Enter the existing workbook which contains the host data-> ")
    fileObject = load_workbook(filename = fileName)
    ReadHostAddrs(fileObject)
    fileName = "Resolved_"+fileName
    fileObject.save(fileName)
    print("[*] DNS Resolution Complete")

if(__name__ == '__main__'):
    main()
    
